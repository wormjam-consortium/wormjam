#!/usr/bin/env python

"""SBTabReader.py - converts a csv with WormBase RNAi identifiers in the first column to a list of corresponding target genes

Requires (use pip to install):
openpyxl
"""

import os
import openpyxl

__author__ = "Jake Hattwell"
__copyright__ = "None"
__credits__ = ["Jake Hattwell"]
__license__ = "CCO"
__version__ = "1"
__maintainer__ = "Jake Hattwell"
__email__ = "j.hattwell@uq.edu.au"
__status__ = "Live"


class modelSystem():
    """Class for use with SBTabInterface.py
    """
    def __init__(self,master=None):
        self.tables = {}
        self.master=master
        self.size = {}
    
    def loadTable(self,name,location):
        self.tables[name] = dataset(location)
        self.size[name] = self.tables[name].rows-2

    def loadFolder(self,name):
        success = False
        self.master.printOut("------------------------")
        if os.path.isdir(name) == False:
            self.master.printOut("This folder does not exist")

        else:
            self.master.printOut("Folder loaded")
            paths = []
            for f in os.listdir(name):
                if "SBtab.xlsx" in f:
                    if name[1] != ":":
                        path = os.getcwd()+"\\"+name+"\\"+f
                    else:
                        path = name+"\\"+f
                        filename = f.replace("-SBtab.xlsx","")
                        paths.append([filename,path])
            if paths == []:
                self.master.printOut(" ".join(["There were no SBtab.xlsx files found in",name]))
            else:
                self.master.printOut("SBtab.xlsx files found! Loading now!")
                for hit in paths:
                    self.master.printOut(" ".join(["Loading file:",hit[0]]))
                    self.loadTable(hit[0],hit[1])
                self.master.printOut(" ".join([str(len(paths)),"files loaded into the model"]))
                success = True    
            
        return success
                    

    def searchModel(self,term,dataset="All"):
        results = {}
        count = 0
        for table,contents in self.tables.items():
            for ID,entry in contents.data.items():
                for key,val in entry.items():
                    try:
                        if term.lower() in str(val).lower():
                            row = list(contents.data).index(ID) + 3
                            results["-".join([table,ID])] = [table,ID,key,str(val),str(row),entry]
                            count += 1
                    except:
                        self.master.printOut("Error searching for term")
        if count != 0:
            self.master.printOut("------------------------")
            
            self.master.printOut(" ".join([str(len(results)),"hits found!"]))
        else:
            self.master.printOut(" ".join(["Search term",term,"returned 0 results"]))
        return results
    
    def prettyPrint(self,accession):
        data = self.tables[accession[0]].data[accession[1]]
        output = ""
        output += " ".join([accession[0]+":",accession[1],"\n"])
        lineCount = 0
        for iden,entry in data.items():
            if iden != None and lineCount < 5 and lineCount > 0:
                output += " ".join([str(iden)+":",str(entry),"\n"])
            lineCount += 1
                
        return output
    

class dataset:
    """Importable class for loading SBTab files\nConverts SBTab as nested dictionary.\n

    instance.data = Dictionary of entries in SBTab\n
    Each entry is a dictionary of the data associated with that entry, with column headers as keys.
        
        Arguments:
            xlsx {str} -- Path to SBTab file of interest.
        
        Keyword Arguments:
            headerRow {int} -- Excel row of the header information, (default: {2})
        """

    def __init__(self,xlsx,headerRow=2):
        """Loads the SBTab file"""
        self.name = xlsx
        self.wb = openpyxl.load_workbook(xlsx)
        self.sheet = self.wb.active
        self.cols = self.sheet.max_column
        self.rows = self.sheet.max_row
        self.sbString = self.sheet.cell(row=1,column=1).value
        self.headerRow = headerRow
        self.headers = [self.sheet.cell(row=2,column = i).value for i in range(1,self.cols+1) if self.sheet.cell(row=self.headerRow,column = i)!= None]
        self.data = {str(self.sheet.cell(row=i,column = 1).value):{self.headers[j-1]:self.sheet.cell(row=i,column=j).value for j in range(1,self.cols+1)}for i in range(self.headerRow+1,self.rows+1)}
        self.freeze_panes = self.sheet.freeze_panes

    def saveToExcel(self,name):
        newWb = openpyxl.Workbook()
        newWs = newWb.active
        newWs['A1'] = self.sbString
        for i in range(len(self.headers)):
            newWs.cell(row=2,column=i+1).value = self.headers[i]
        row = 3
        for key,val in self.data.items():
            del key
            col = 1
            for i in self.headers:
                newWs.cell(row=row,column=col).value = val[i]
                col += 1
            row += 1

        newWs.freeze_panes = self.freeze_panes
        newWb.save(name+'-SBtab.xlsx')
